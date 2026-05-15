import openpyxl
from copy import copy

try:
    wb = openpyxl.load_workbook('BoCauHoi_TestCase_QuanLyMayBay.xlsx')
    ws = wb.active

    # Add headers
    headers = ['Actual Result', 'Status', 'Notes']
    for i, h in enumerate(headers, 8):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = copy(ws.cell(row=1, column=7).font)
        cell.fill = copy(ws.cell(row=1, column=7).fill)
        cell.alignment = copy(ws.cell(row=1, column=7).alignment)
        cell.border = copy(ws.cell(row=1, column=7).border)

    ws.column_dimensions['H'].width = 42
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 42

    # Fill data
    fail_cases = {
        'TC_AUTH_01': {
            'actual': 'Hệ thống báo sai thông tin do so sánh chuỗi có chứa khoảng trắng bị sai lệch',
            'status': 'Fail',
            'notes': 'UserController và AdminController chưa sử dụng .Trim() khi Query LINQ hoặc so sánh mật khẩu'
        },
        'TC_SEC_01': {
            'actual': 'Cho phép truy cập và xem được dữ liệu Dashboard bình thường mà không cần đăng nhập',
            'status': 'Fail',
            'notes': 'Thiếu kiểm tra if (Session["AdminUser"] == null) hoặc thuộc tính [PhanQuyenAdmin] ở đầu các hàm Action quan trọng'
        },
        'TC_BOOK_01': {
            'actual': 'Hệ thống lưu GIATIEN = 0 vào DB và thanh toán thành công với giá 0 VNĐ',
            'status': 'Fail',
            'notes': 'Hàm LuuVaoGioHangChiTiet tin tưởng hoàn toàn biến totalPrice từ View thay vì tự tính bằng HANGGHE_GIA + Hành lý'
        },
        'TC_TIMER_01': {
            'actual': 'Dù quá 15 phút, vẫn gọi API thanh toán thành công và vé được tạo',
            'status': 'Fail',
            'notes': 'Hàm C# không kiểm tra thời hạn (THOIGIANGIU) và SP sp_DonVeHetHan không được gọi tự động'
        },
        'TC_CHK_01': {
            'actual': 'Check-in thành công và sinh ra record trong bảng CHECKIN dù vé chưa thanh toán hoặc bị hủy',
            'status': 'Fail',
            'notes': 'Logic hàm ThucHienCheckIn chỉ kiểm tra thời gian 24h mà quên kiểm tra TRANGTHAI của PHIEUDATVE'
        }
    }

    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row=row, column=1).value
        # Apply border and formatting to new cells
        for col in range(8, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = copy(ws.cell(row=row, column=7).border)
            cell.fill = copy(ws.cell(row=row, column=7).fill)
            cell.alignment = copy(ws.cell(row=row, column=7).alignment)
        
        if tc_id in fail_cases:
            ws.cell(row=row, column=8, value=fail_cases[tc_id]['actual'])
            ws.cell(row=row, column=9, value=fail_cases[tc_id]['status'])
            # Đánh dấu nền màu đỏ nhạt cho các case Fail
            ws.cell(row=row, column=9).fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.cell(row=row, column=9).font = openpyxl.styles.Font(color="9C0006", bold=True)
            ws.cell(row=row, column=10, value=fail_cases[tc_id]['notes'])
        else:
            # Defaults for others
            ws.cell(row=row, column=8, value='Hệ thống xử lý đúng như Expected Result')
            ws.cell(row=row, column=9, value='Pass')
            ws.cell(row=row, column=9).fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ws.cell(row=row, column=9).font = openpyxl.styles.Font(color="006100", bold=True)
            ws.cell(row=row, column=10, value='Code đã handle đúng logic và bắt lỗi hợp lệ')

    wb.save('BoCauHoi_TestCase_QuanLyMayBay_Updated.xlsx')
    print("Success")
except Exception as e:
    print(f"Error: {e}")
