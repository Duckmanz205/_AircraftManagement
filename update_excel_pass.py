import openpyxl

try:
    wb = openpyxl.load_workbook('BoCauHoi_TestCase_QuanLyMayBay_Updated.xlsx')
    ws = wb.active

    # Các Test Case đã được fix code ở Controller
    fixed_cases = {
        'TC_AUTH_01': {
            'actual': 'Đã sử dụng hàm .Trim() để loại bỏ các khoảng trắng thừa. Hệ thống xử lý đăng nhập thành công như mong đợi.',
            'status': 'Pass',
            'notes': 'Đã bổ sung .Trim() vào UserController và AdminController.'
        },
        'TC_SEC_01': {
            'actual': 'Hệ thống phát hiện Session rỗng và điều hướng về trang Đăng nhập.',
            'status': 'Pass',
            'notes': 'Đã bổ sung kiểm tra if (Session["AdminUser"] == null) vào đầu tất cả các hàm Action quan trọng trong AdminController.'
        },
        'TC_BOOK_01': {
            'actual': 'Hệ thống tự động query HANGGHE_GIA từ CSDL để tính tổng tiền, từ chối mức giá 0 VNĐ bị can thiệp.',
            'status': 'Pass',
            'notes': 'Đã xóa việc lấy model.totalPrice từ Frontend. Backend đã có logic tự tính toán giá vé an toàn.'
        },
        'TC_TIMER_01': {
            'actual': 'Hệ thống chặn lại và báo lỗi Giỏ hàng đã hết hạn giữ chỗ (Quá 15 phút), xóa thông tin giữ ghế.',
            'status': 'Pass',
            'notes': 'Đã chèn check THOIGIANGIU < DateTime.Now và gọi SP sp_DonVeHetHan vào hàm thanh toán.'
        },
        'TC_CHK_01': {
            'actual': 'Hệ thống thông báo "Vé không hợp lệ, chưa thanh toán hoặc đã bị hủy" và chặn thao tác Check-in.',
            'status': 'Pass',
            'notes': 'Hàm ThucHienCheckIn đã bổ sung điều kiện kiểm tra TRANGTHAI == "Đã thanh toán" cho PhieuDatVe.'
        }
    }

    # Định dạng Pass (Màu xanh)
    pass_fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pass_font = openpyxl.styles.Font(color="006100", bold=True)

    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row=row, column=1).value
        if tc_id in fixed_cases:
            # Update values
            ws.cell(row=row, column=8, value=fixed_cases[tc_id]['actual'])
            ws.cell(row=row, column=9, value=fixed_cases[tc_id]['status'])
            ws.cell(row=row, column=10, value=fixed_cases[tc_id]['notes'])
            
            # Apply format
            ws.cell(row=row, column=9).fill = pass_fill
            ws.cell(row=row, column=9).font = pass_font

    wb.save('BoCauHoi_TestCase_QuanLyMayBay_Updated.xlsx')
    print("Success")
except Exception as e:
    print(f"Error: {e}")
