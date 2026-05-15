import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Tạo workbook mới
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bug Report"

# Định nghĩa header
headers = ["Test Case ID", "Vấn đề phát hiện", "Vị trí lỗi", "Code sửa đổi (Fix)", "Giải thích"]

# Ghi header
for col_num, header_title in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header_title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Border
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border

# Dữ liệu các lỗi tìm thấy
data = [
    [
        "TC_AUTH_01",
        "Lỗi đăng nhập: Không Trim() khoảng trắng ở trường EMAIL và MATKHAU, dẫn đến người dùng nhập dư khoảng trắng hoặc dữ liệu trong SQL bị dư khoảng trắng sẽ không khớp.",
        "UserController.cs (hàm DangNhapThanhCong, dòng 79)\nAdminController.cs (hàm DangNhapThanhCong, dòng 34)",
        "// UserController.cs\nKHACHHANG kh = db.KHACHHANGs.FirstOrDefault(x => x.EMAIL.Trim() == Email.Trim());\nif (kh.MATKHAU.Trim() != password.Trim())\n\n// AdminController.cs\nNHANVIEN nhanVien = db.NHANVIENs.FirstOrDefault(nv => nv.MANV.Trim() == username.Trim());\nif (nhanVien.MATKHAU.Trim() != password.Trim())",
        "Sử dụng .Trim() cho cả hai vế so sánh để đảm bảo tính chính xác, loại bỏ các khoảng trắng thừa từ đầu vào UI hoặc từ đặc thù của kiểu dữ liệu chuỗi trong SQL Server."
    ],
    [
        "TC_AUTH_02",
        "Xử lý Session khi login thất bại bị lỗi. Nếu nhập sai mật khẩu (set Session ErrorPassword), sau đó nhập sai email ở lần tiếp theo, Session ErrorPassword cũ không bị xóa đi, dẫn đến UI hiển thị sai thông báo lỗi.",
        "UserController.cs (hàm DangNhapThanhCong, dòng 76)",
        "// Thêm vào đầu hàm:\nSession[\"ErrorEmail\"] = null;\nSession[\"ErrorPassword\"] = null;\n\n// Sau đó mới thực hiện logic if (kh == null) ...",
        "Luôn phải reset trạng thái Session lỗi mỗi khi bắt đầu xử lý request đăng nhập mới để đảm bảo các thông báo hiển thị trên View đúng với lỗi hiện tại."
    ],
    [
        "TC_BOOK_01",
        "Lỗ hổng bảo mật thanh toán: Module Đặt vé lưu GIATIEN bằng cách tin tưởng hoàn toàn vào biến model.totalPrice gửi lên từ Frontend. Hacker có thể can thiệp request để mua vé với giá 0đ.",
        "UserController.cs (hàm LuuVaoGioHangChiTiet, dòng 621)",
        "// Cần lấy giá trị từ DB thay vì model.totalPrice\nvar hangGhe = db.HANGGHE_GIA.FirstOrDefault(x => x.MACB == model.MaCB && x.HANGGHE == model.Passengers.FirstOrDefault().SeatClass);\nvar giaCoSo = hangGhe != null ? hangGhe.GIA_COSO : 0;\nvar tongHanhLy = model.Passengers.Sum(p => p.CarryOnFee + p.CheckedFee);\nchiTiet.GIATIEN = (giaCoSo * model.Passengers.Count) + tongHanhLy;\n\n// Trong SQL Server cũng nên sử dụng FN_TinhTongTien để đối soát lại trước khi insert.",
        "Không bao giờ được tin tưởng dữ liệu tính toán tiền tệ từ Client. Backend bắt buộc phải tính toán lại giá vé và phụ phí hành lý từ dữ liệu gốc trong Database."
    ],
    [
        "TC_SEC_01",
        "Lỗ hổng phân quyền (Missing Authorization). Các Action quan trọng trong quản trị như TrangChu, ThongKe, QLChuyenBay không có thuộc tính [PhanQuyenAdmin] hoặc kiểm tra Session. Ai biết URL cũng có thể truy cập trái phép.",
        "AdminController.cs (hàm TrangChu, ThongKe, QLChuyenBay, v.v...)",
        "// Có thể thêm vào đầu mỗi hàm hoặc dùng Action Filter:\nif (Session[\"AdminUser\"] == null) {\n    return RedirectToAction(\"DangNhap\", \"User\");\n}\n\n// Hoặc gắn thêm annotation cho các Action cần thiết:\n[PhanQuyenAdmin(\"CV09\", \"CV05\")]\npublic ActionResult ThongKe(...)",
        "Đảm bảo nguyên tắc 'Zero Trust' - mọi endpoint dành cho quản trị viên đều phải trải qua bước xác thực (Authentication) và phân quyền (Authorization) ngay tại đầu Action."
    ],
    [
        "TC_CHK_01",
        "Lỗi logic Check-in: Bỏ qua việc kiểm tra trạng thái thanh toán của vé. Một vé chưa thanh toán hoặc đã bị hủy (TRANGTHAI = 'Hủy') vẫn có thể vượt qua bước kiểm tra và Check-in thành công.",
        "UserController.cs (hàm ThucHienCheckIn, dòng 1479)",
        "// Sau khi tìm thấy phieuDatVe, thêm kiểm tra trạng thái:\nif (phieuDatVe == null) \n    return Json(new { success = false, message = \"Vé không thuộc về bạn\" });\n\nif (phieuDatVe.TRANGTHAI != \"Đã thanh toán\")\n    return Json(new { success = false, message = \"Vé chưa được thanh toán hoặc đã bị hủy.\" });",
        "Chức năng Check-in bắt buộc phải diễn ra trên các vé hợp lệ, nghĩa là giao dịch mua vé đã được hoàn tất và ghi nhận trạng thái 'Đã thanh toán'."
    ]
]

# Ghi dữ liệu
for r_idx, row_data in enumerate(data, 2):
    for c_idx, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

# Định dạng độ rộng cột
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 60
ws.column_dimensions['E'].width = 40

# Lưu file
file_name = "Bug_Report_FlightManager.xlsx"
wb.save(file_name)
print(f"Excel file '{file_name}' generated successfully.")
