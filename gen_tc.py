import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

headers = ["Test Case#","Test Title/Scenario","Test Summary","Test Steps","Test Data","Expected Result","Post-condition"]
col_widths = [15,30,35,55,40,40,40]

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

for i,h in enumerate(headers,1):
    cell = ws.cell(row=1,column=i,value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = border
    ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

ws.row_dimensions[1].height = 30
